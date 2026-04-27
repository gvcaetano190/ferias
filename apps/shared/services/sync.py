from __future__ import annotations

import csv
import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests
import certifi
from django.conf import settings

from apps.core.models import OperationalSettings
from apps.block.models import BlockProcessing
from apps.notifications.services import NotificationService
from apps.people.models import Acesso, Ferias
from apps.shared.repositories.people import AcessoRepository, ColaboradorRepository, FeriasRepository
from apps.shared.repositories.sync import SyncLogRepository
from apps.shared.services.google_sheets import build_export_url, extract_sheet_id


class SpreadsheetSyncService:
    MAX_EXPECTED_VACATION_DAYS = 62
    DELETE_BATCH_SIZE = 500
    OPERATIONAL_SYSTEMS = {"AD PRIN", "VPN"}
    # Status "fraco" = incerto, sem informação → preserva o que o AD disse
    # NB = campo vazio/não mapeado na planilha
    UNCERTAIN_ACCESS_STATUSES = {"", "-", "NB"}

    # NP = "Não Possui" → definitivo, a pessoa não tem essa ferramenta
    # BLOQUEADO/LIBERADO → definitivo, sobrescreve sempre
    AUTHORITATIVE_ACCESS_STATUSES = {"BLOQUEADO", "BLOQUEADA", "LIBERADO", "LIBERADA"}

    # Manter compatibilidade com código legado que usa WEAK_ACCESS_STATUSES
    WEAK_ACCESS_STATUSES = {"", "-", "NB", "NP"}

    def resolve_access_status(
        self,
        *,
        collaborator_id: int,
        system_name: str,
        imported_status: str,
    ) -> str:
        imported_normalized = (imported_status or "").strip().upper()

        # Sistemas não operacionais: aceita qualquer coisa da planilha
        if system_name not in self.OPERATIONAL_SYSTEMS:
            return imported_status

        # NP = "Não Possui": definitivo, a pessoa não tem essa ferramenta.
        # Sobrescreve sempre, sem preservar histórico do AD.
        if imported_normalized == "NP":
            return "NP"

        # BLOQUEADO / LIBERADO: status explícito da planilha, aceita sempre.
        if imported_normalized in self.AUTHORITATIVE_ACCESS_STATUSES:
            return imported_status

        # NB / vazio: status incerto → preserva o que o AD registrou anteriormente.
        if imported_normalized in self.UNCERTAIN_ACCESS_STATUSES:
            authoritative_status = self._latest_authoritative_status(
                collaborator_id=collaborator_id,
                system_name=system_name,
            )
            if authoritative_status in self.AUTHORITATIVE_ACCESS_STATUSES:
                return authoritative_status

        return imported_status

    def __init__(self):
        self.operational_settings = OperationalSettings.get_solo()
        self.colaboradores = ColaboradorRepository()
        self.acessos = AcessoRepository()
        self.sync_logs = SyncLogRepository()
        self.notification_service = NotificationService()
        settings.DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
        settings.PENDING_SYNC_CSV.parent.mkdir(parents=True, exist_ok=True)

    def run(self, force: bool = False) -> dict[str, Any]:
        if not self.operational_settings.sync_enabled:
            payload = {"status": "disabled", "message": "Sincronizacao desabilitada no admin."}
            self._notify_sync_status(payload, force=force)
            return payload

        try:
            spreadsheet = self.download_spreadsheet(force=force)
            file_hash = self.calculate_hash(spreadsheet)
            if not force and file_hash == self.last_hash():
                payload = {"status": "skipped", "message": "A planilha nao mudou desde a ultima sincronizacao."}
                self._notify_sync_status(payload, force=force)
                return payload

            records, sheets = self.process_workbook(spreadsheet)

            pending: list[dict[str, str]] = []
            total_events = 0
            total_accesses = 0
            seen_ferias_keys: set[tuple[int, str, str, int, int]] = set()
            seen_access_keys: set[tuple[int, str]] = set()

            for record in records:
                collaborator = self.resolve_collaborator(record)
                if not collaborator:
                    pending.append(
                        {
                            "nome": record["nome"],
                            "email": record["email"],
                            "login_ad": record["login_ad"],
                            "aba": record["aba_origem"],
                            "data_saida": record["data_saida"],
                            "data_retorno": record["data_retorno"],
                        }
                    )
                    continue

                seen_ferias_keys.add(
                    (
                        collaborator.id,
                        record["data_saida"],
                        record["data_retorno"],
                        record["mes"],
                        record["ano"],
                    )
                )
                Ferias.objects.update_or_create(
                    colaborador_id=collaborator.id,
                    data_saida=record["data_saida"],
                    data_retorno=record["data_retorno"],
                    mes_ref=record["mes"],
                    ano_ref=record["ano"],
                    defaults={},
                )
                total_events += 1

                for system_name, status in record["acessos"].items():
                    seen_access_keys.add((collaborator.id, system_name))
                    self.acessos.upsert(
                        colaborador_id=collaborator.id,
                        sistema=system_name,
                        status=self.resolve_access_status(
                            collaborator_id=collaborator.id,
                            system_name=system_name,
                            imported_status=status,
                        ),
                    )
                    total_accesses += 1

            self.reconcile_operational_sync_data(
                seen_ferias_keys=seen_ferias_keys,
                seen_access_keys=seen_access_keys,
            )

            if pending:
                self.write_pending_csv(pending)
            elif settings.PENDING_SYNC_CSV.exists():
                settings.PENDING_SYNC_CSV.unlink()

            sync_status = "SUCCESS" if not pending else "PARTIAL"
            self.sync_logs.create(
                tipo_sync="django_planilha",
                status=sync_status,
                total_registros=total_events + total_accesses,
                total_abas=len(sheets),
                mensagem=(
                    f"Sincronizacao via Django: ferias={total_events}, "
                    f"acessos={total_accesses}, pendencias={len(pending)}"
                ),
                arquivo_hash=file_hash,
                detalhes=f"arquivo={spreadsheet.name}; pendencias={len(pending)}",
            )

            payload = {
                "status": "success",
                "message": f"Sincronizados {total_events} eventos e {total_accesses} acessos.",
                "records": total_events + total_accesses,
                "sheets": len(sheets),
                "pending": len(pending),
                "file": str(spreadsheet),
            }
            self._notify_sync_status(payload, force=force)
            return payload
        except Exception as exc:
            payload = {"status": "error", "message": str(exc)}
            self._notify_sync_status(payload, force=force)
            raise

    def _notify_sync_status(self, payload: dict[str, Any], *, force: bool) -> None:
        self.notification_service.notify_task_status(
            task_key="spreadsheet_sync",
            task_label="Sincronizacao da planilha",
            status=payload.get("status", "unknown"),
            summary=payload.get("message", "Rotina finalizada."),
            details=[
                f"Modo forcado: {'sim' if force else 'nao'}",
                f"Registros: {payload.get('records', 0)}",
                f"Abas: {payload.get('sheets', 0)}",
                f"Pendencias: {payload.get('pending', 0)}",
                f"Arquivo: {payload.get('file', '-')}",
            ],
        )

    def reconcile_operational_sync_data(
        self,
        *,
        seen_ferias_keys: set[tuple[int, str, str, int, int]],
        seen_access_keys: set[tuple[int, str]],
    ) -> None:
        stale_ferias_ids: list[int] = []
        for ferias_id, colaborador_id, data_saida, data_retorno, mes_ref, ano_ref in Ferias.objects.values_list(
            "id",
            "colaborador_id",
            "data_saida",
            "data_retorno",
            "mes_ref",
            "ano_ref",
        ):
            current_key = (
                colaborador_id,
                self._normalize_key_part(data_saida),
                self._normalize_key_part(data_retorno),
                mes_ref,
                ano_ref,
            )
            if current_key not in seen_ferias_keys:
                stale_ferias_ids.append(ferias_id)

        self._delete_in_batches(Ferias, stale_ferias_ids)

        stale_access_ids: list[int] = []
        for acesso_id, colaborador_id, sistema in Acesso.objects.values_list("id", "colaborador_id", "sistema"):
            current_key = (colaborador_id, sistema)
            if current_key not in seen_access_keys:
                stale_access_ids.append(acesso_id)

        self._delete_in_batches(Acesso, stale_access_ids)


    def _latest_authoritative_status(self, *, collaborator_id: int, system_name: str) -> str:
        latest_processing = (
            BlockProcessing.objects.filter(
                colaborador_id=collaborator_id,
                resultado__in={"SUCESSO", "SINCRONIZADO"},
            )
            .order_by("-executado_em", "-id")
            .first()
        )
        if latest_processing:
            if system_name == "AD PRIN":
                status = (latest_processing.ad_status or "").strip().upper()
                if status in self.AUTHORITATIVE_ACCESS_STATUSES:
                    return status
            if system_name == "VPN":
                status = (latest_processing.vpn_status or "").strip().upper()
                if status in self.AUTHORITATIVE_ACCESS_STATUSES:
                    return status

        current_access = (
            Acesso.objects.filter(colaborador_id=collaborator_id, sistema=system_name)
            .order_by("-updated_at", "-id")
            .first()
        )
        return ((getattr(current_access, "status", "") or "").strip().upper())

    def _delete_in_batches(self, model, stale_ids: list[int]) -> None:
        if not stale_ids:
            return
        for index in range(0, len(stale_ids), self.DELETE_BATCH_SIZE):
            batch = stale_ids[index : index + self.DELETE_BATCH_SIZE]
            model.objects.filter(id__in=batch).delete()

    def _normalize_key_part(self, value):
        if value is None:
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return value

    def resolve_collaborator(self, record: dict[str, Any]):
        from apps.people.models import Colaborador

        email = (record.get("email") or "").strip().lower()
        login_ad = (record.get("login_ad") or "").strip().lower()
        nome = (record.get("nome") or "").strip()

        if email:
            collaborator = Colaborador.objects.filter(email__iexact=email).first()
            if collaborator:
                return collaborator
        if login_ad:
            collaborator = Colaborador.objects.filter(login_ad__iexact=login_ad).first()
            if collaborator:
                return collaborator
        return Colaborador.objects.filter(nome=nome).first()

    def download_spreadsheet(self, force: bool = False) -> Path:
        cached_files = sorted(
            settings.DOWNLOAD_DIR.glob("planilha_*.xlsx"),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        if cached_files and not force:
            latest = cached_files[0]
            age_minutes = (datetime.now().timestamp() - latest.stat().st_mtime) / 60
            if age_minutes <= self.operational_settings.cache_minutes:
                return latest

        url = self.operational_settings.google_sheets_url or settings.GOOGLE_SHEETS_URL
        sheet_id = extract_sheet_id(url)
        if not sheet_id:
            raise ValueError("Google Sheets URL inválida.")

        download_url = build_export_url(sheet_id, file_format="xlsx")
        target = settings.DOWNLOAD_DIR / f"planilha_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.fetch_file(download_url, target)
        return target

    def fetch_file(self, url: str, target: Path) -> None:
        try:
            response = requests.get(url, timeout=60, stream=True, verify=certifi.where())
            response.raise_for_status()
        except requests.RequestException as exc:
            raise ValueError(f"Falha ao baixar planilha do Google Sheets: {exc}") from exc

        with target.open("wb") as file_handle:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    file_handle.write(chunk)

    def process_workbook(self, spreadsheet: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        workbook = openpyxl.load_workbook(spreadsheet, data_only=True)
        records: list[dict[str, Any]] = []
        processed_sheets: list[dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            month, year = self.extract_month_year(sheet_name)
            rows = self.process_sheet(workbook[sheet_name], sheet_name, month, year)
            records.extend(rows)
            processed_sheets.append(
                {
                    "nome": sheet_name,
                    "mes": month,
                    "ano": year,
                    "total_funcionarios": len(rows),
                }
            )

        return records, processed_sheets

    def process_sheet(self, worksheet, sheet_name: str, month: int, year: int) -> list[dict[str, Any]]:
        header_indexes: dict[int, str] = {}
        header_lookup: dict[str, int] = {}
        for index, cell in enumerate(worksheet[1], start=0):
            if cell.value is None:
                continue
            column_name = str(cell.value).strip()
            header_indexes[index] = column_name
            header_lookup[column_name.upper()] = index

        indexes = {
            "unidade": self.first_match(header_lookup, ["RESP.", "RESP", "UNIDADE", "RESPONSAVEL", "RESPONSÁVEL"], 0),
            "nome": self.first_match(header_lookup, ["NOME", "FUNCIONÁRIO", "FUNCIONARIO", "COLABORADOR"], 1),
            "email": self.first_match(header_lookup, ["EMAIL", "E-MAIL", "MAIL"]),
            "motivo": self.first_match(header_lookup, ["MOTIVO", "TIPO", "RAZÃO", "RAZAO"], 2),
            "saida": self.first_match(header_lookup, ["SAÍDA", "SAIDA", "DATA SAÍDA", "DATA SAIDA", "INÍCIO", "INICIO"], 3),
            "retorno": self.first_match(header_lookup, ["RETORNO", "RETORNO/LIBERAÇÃO", "RETORNO/LIBERACAO", "LIBERAÇÃO", "LIBERACAO", "DATA RETORNO", "FIM"], 4),
            "gestor": next((idx for name, idx in header_lookup.items() if "GESTOR" in name), None),
        }

        system_indexes: dict[str, int] = {}
        for idx, column_name in header_indexes.items():
            name_upper = column_name.upper()
            for system_name in self.allowed_systems:
                if system_name.upper() in name_upper:
                    system_indexes[system_name] = idx
                    break

        rows: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=2):
            raw_name = self.value_from_row(row, indexes["nome"])
            if self.is_blank(raw_name):
                continue

            raw_email = self.value_from_row(row, indexes["email"])
            raw_start = self.value_from_row(row, indexes["saida"])
            raw_end = self.value_from_row(row, indexes["retorno"])
            start_date = self.normalize_date(self.parse_date(raw_start), month, year)
            end_date = self.normalize_return_date(self.parse_date(raw_end), start_date)
            if not start_date or not end_date:
                continue

            email = self.normalize_email(raw_email)
            row_data = {
                "nome": str(raw_name).strip(),
                "email": email,
                "login_ad": self.extract_login(email),
                "unidade": self.clean_text(self.value_from_row(row, indexes["unidade"])),
                "motivo": self.clean_text(self.value_from_row(row, indexes["motivo"])),
                "data_saida": start_date.strftime("%Y-%m-%d"),
                "data_retorno": end_date.strftime("%Y-%m-%d"),
                "gestor": self.clean_text(self.value_from_row(row, indexes["gestor"])),
                "aba_origem": sheet_name,
                "mes": month,
                "ano": year,
                "acessos": {},
            }

            for system_name in self.allowed_systems:
                status = self.map_access_status(self.value_from_row(row, system_indexes.get(system_name)))
                row_data["acessos"][system_name] = status
            rows.append(row_data)

        return rows

    @property
    def allowed_systems(self) -> list[str]:
        return self.operational_settings.allowed_systems or list(settings.DEFAULT_ACCESS_SYSTEMS)

    def last_hash(self) -> str:
        latest_log = self.sync_logs.latest_by_type("django_planilha")
        if latest_log and latest_log.arquivo_hash:
            return str(latest_log.arquivo_hash)
        return ""

    def calculate_hash(self, path: Path) -> str:
        with path.open("rb") as file_handle:
            return hashlib.md5(file_handle.read()).hexdigest()

    def write_pending_csv(self, rows: list[dict[str, str]]) -> None:
        with settings.PENDING_SYNC_CSV.open("w", encoding="utf-8", newline="") as file_handle:
            writer = csv.DictWriter(file_handle, fieldnames=["nome", "email", "login_ad", "aba", "data_saida", "data_retorno"])
            writer.writeheader()
            writer.writerows(rows)

    def extract_month_year(self, sheet_name: str) -> tuple[int, int]:
        months = {
            "JANEIRO": 1,
            "FEVEREIRO": 2,
            "MARÇO": 3,
            "MARCO": 3,
            "ABRIL": 4,
            "MAIO": 5,
            "JUNHO": 6,
            "JULHO": 7,
            "AGOSTO": 8,
            "SETEMBRO": 9,
            "OUTUBRO": 10,
            "NOVEMBRO": 11,
            "DEZEMBRO": 12,
        }
        upper_name = sheet_name.upper()
        month = next((value for key, value in months.items() if key in upper_name), datetime.now().month)

        four_digits = re.search(r"(20\d{2})", sheet_name)
        if four_digits:
            return month, int(four_digits.group(1))

        two_digits = re.search(r"(\d{2})$", sheet_name)
        if two_digits:
            year = int(two_digits.group(1))
            return month, 2000 + year if year <= 30 else 1900 + year

        return month, datetime.now().year

    def first_match(self, lookup: dict[str, int], candidates: list[str], fallback: int | None = None) -> int | None:
        for candidate in candidates:
            if candidate in lookup:
                return lookup[candidate]
        return fallback

    def value_from_row(self, row, index: int | None):
        if index is None or index >= len(row):
            return None
        return row[index].value

    def clean_text(self, value: Any) -> str:
        if self.is_blank(value):
            return ""
        return str(value).strip()

    def is_blank(self, value: Any) -> bool:
        if value is None:
            return True
        string_value = str(value).strip().lower()
        return string_value in {"", "nan", "none"}

    def parse_date(self, value: Any) -> date | None:
        if self.is_blank(value):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(value).split()[0], fmt).date()
            except ValueError:
                continue
        return None

    def normalize_date(self, value: date | None, month: int, year: int) -> date | None:
        if value is None:
            return value

        normalized = value
        if normalized.day <= 12 and normalized.month != month:
            try:
                inverted = date(year, normalized.day, normalized.month)
            except ValueError:
                inverted = normalized
            if inverted.month == month:
                normalized = inverted

        if normalized.month == month and normalized.year != year:
            try:
                normalized = normalized.replace(year=year)
            except ValueError:
                return normalized

        return normalized

    def normalize_return_date(self, return_date: date | None, start_date: date | None) -> date | None:
        if return_date is None or start_date is None:
            return return_date

        normalized = return_date
        inverted = None
        if normalized.day <= 12:
            try:
                inverted = date(normalized.year, normalized.day, normalized.month)
            except ValueError:
                inverted = None

        if inverted and normalized.month != start_date.month:
            # Corrige casos como 04/10 interpretado como outubro quando o contexto real é 10/04.
            if inverted.month == start_date.month and inverted >= start_date:
                normalized = inverted

        if normalized < start_date and normalized.day <= 12:
            try:
                inverted = date(normalized.year, normalized.day, normalized.month)
            except ValueError:
                inverted = normalized
            if inverted >= start_date:
                normalized = inverted

        if normalized.year != start_date.year:
            try:
                same_year = normalized.replace(year=start_date.year)
            except ValueError:
                same_year = normalized
            if same_year >= start_date:
                normalized = same_year
            else:
                try:
                    next_year = normalized.replace(year=start_date.year + 1)
                except ValueError:
                    next_year = normalized
                if next_year >= start_date:
                    normalized = next_year

        if inverted and inverted >= start_date:
            current_duration = (normalized - start_date).days
            inverted_duration = (inverted - start_date).days
            # Se a data atual gera um afastamento muito longo e a invertida gera uma duração plausível,
            # preferimos a invertida para evitar casos como 04/10 no lugar de 10/04.
            if current_duration > self.MAX_EXPECTED_VACATION_DAYS and 0 <= inverted_duration <= self.MAX_EXPECTED_VACATION_DAYS:
                normalized = inverted

        return normalized

    def normalize_email(self, value: Any) -> str:
        if self.is_blank(value):
            return ""
        email = str(value).strip().lower()
        return "" if email in {"#n/d", "#n/a", "np", "n/a"} else email

    def extract_login(self, email: str) -> str:
        if not email:
            return ""
        suffix = "@printi.com.br"
        if email.endswith(suffix):
            return email[:-len(suffix)]
        return email.split("@")[0]

    def map_access_status(self, value: Any) -> str:
        if str(value or "").strip() == "-":
            return "NP"
        if self.is_blank(value):
            return "NB"
        normalized = str(value).strip().upper()
        if normalized in {"BLOQUEADO", "BLOQ"}:
            return "BLOQUEADO"
        if normalized in {"LIBERADO", "LIB"}:
            return "LIBERADO"
        if normalized in {"-", "NP", "N/P", "N\\A", "NA", "N/A"}:
            return "NP"
        return "NB"
